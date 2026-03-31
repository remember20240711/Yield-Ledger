from io import BytesIO
from datetime import datetime, timezone
from decimal import Decimal
from urllib.parse import quote

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import delete, or_, select
from sqlalchemy.orm import Session, selectinload

from app.db.session import SessionLocal, get_db
from app.models.dividend import Dividend
from app.models.stock import Stock
from app.models.transaction import Transaction
from app.schemas.stock import (
    CreateStockRequest,
    CreateTransactionRequest,
    DividendDetailResponse,
    PortfolioExportResponse,
    PortfolioImportRequest,
    PortfolioImportResponse,
    PortfolioSummaryResponse,
    StockSearchResponse,
    StockListResponse,
    StockRowResponse,
    TransactionDetailResponse,
)
from app.services.catalog import stock_catalog_service
from app.services.market import MarketResolver
from app.services.portfolio import (
    serialize_dividends,
    serialize_stock,
    serialize_summary,
    serialize_transactions,
    summarize_transactions,
)
from app.services.provider import market_data_service
from app.utils.numbers import money, shares, trade_price


router = APIRouter(tags=["dividend-tracker"])
market_resolver = MarketResolver()


def sync_stock_cache_task(stock_id: int) -> None:
    # 后台异步刷新缓存，避免新增持仓或交易时阻塞主请求。
    db = SessionLocal()
    try:
        market_data_service.refresh_stock_by_id(db, stock_id)
    finally:
        db.close()


def load_stock_with_relations(db: Session, stock_id: int) -> Stock:
    # 列表和详情页都会用到完整关联数据，统一走这个加载函数。
    stock = db.scalar(
        select(Stock)
        .where(Stock.id == stock_id)
        .options(selectinload(Stock.transactions), selectinload(Stock.dividends))
    )
    if stock is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="股票不存在")
    return stock


def load_portfolio_export_items(db: Session) -> list[dict]:
    # 导出会复用同一份快照结构，保证 JSON 和 Excel 字段一致。
    stocks = db.scalars(
        select(Stock)
        .options(selectinload(Stock.transactions), selectinload(Stock.dividends))
        .order_by(Stock.created_at.asc(), Stock.id.asc())
    ).all()
    items: list[dict] = []
    for stock in stocks:
        items.append(
            {
                "symbol": stock.symbol,
                "normalized_symbol": stock.normalized_symbol,
                "name": stock.name,
                "market": stock.market,
                "currency": stock.currency,
                "last_price": float(stock.last_price) if stock.last_price is not None else None,
                "latest_dividend_ttm": float(stock.latest_dividend_ttm),
                "current_dividend_yield": float(stock.current_dividend_yield),
                "five_year_avg_yield": float(stock.five_year_avg_yield),
                "ten_year_avg_yield": float(stock.ten_year_avg_yield),
                "last_synced_at": stock.last_synced_at,
                "sync_status": stock.sync_status,
                "sync_message": stock.sync_message,
                "transactions": [
                    {
                        "transaction_type": tx.transaction_type,
                        "trade_date": tx.trade_date,
                        "shares": float(tx.shares),
                        "average_price": float(tx.average_price),
                        "total_amount": float(tx.total_amount),
                    }
                    for tx in sorted(stock.transactions, key=lambda item: (item.trade_date, item.id))
                ],
                "dividends": [
                    {
                        "year": item.year,
                        "dividend_per_share": float(item.dividend_per_share),
                        "dividend_yield": float(item.dividend_yield or 0),
                        "close_price": float(item.close_price or 0),
                        "currency": item.currency,
                        "source": item.source,
                    }
                    for item in sorted(stock.dividends, key=lambda dividend: dividend.year)
                ],
            }
        )
    return items


@router.get("/health")
def health() -> dict:
    return {"status": "ok"}


@router.get("/summary", response_model=PortfolioSummaryResponse)
def get_summary(db: Session = Depends(get_db)) -> dict:
    # 顶部汇总卡片直接复用组合序列化逻辑。
    stocks = db.scalars(
        select(Stock).options(selectinload(Stock.transactions), selectinload(Stock.dividends))
    ).all()
    return serialize_summary(stocks)


@router.get("/stocks", response_model=StockListResponse)
def list_stocks(db: Session = Depends(get_db)) -> dict:
    stocks = db.scalars(
        select(Stock)
        .options(selectinload(Stock.transactions), selectinload(Stock.dividends))
        .order_by(Stock.created_at.desc())
    ).all()
    summary = serialize_summary(stocks)
    return {"items": [serialize_stock(stock) for stock in stocks], "base_currency": summary["base_currency"]}


@router.get("/stocks/search", response_model=StockSearchResponse)
def search_stocks(q: str, limit: int = 20, db: Session = Depends(get_db)) -> dict:
    limited = max(1, min(limit, 50))
    return {"items": stock_catalog_service.search(db, q, limited)}


@router.get("/portfolio/export", response_model=PortfolioExportResponse)
def export_portfolio(db: Session = Depends(get_db)) -> dict:
    # 导出完整持仓快照，便于本地备份或迁移到其他实例。
    items = load_portfolio_export_items(db)
    return {
        "version": "1.0",
        "exported_at": datetime.now(timezone.utc),
        "stocks": items,
    }


@router.get("/portfolio/export.xlsx")
def export_portfolio_excel(db: Session = Depends(get_db)) -> StreamingResponse:
    # Excel 导出用于人工查看/分析，保留多 sheet 结构。
    items = load_portfolio_export_items(db)
    stocks = db.scalars(
        select(Stock).options(selectinload(Stock.transactions), selectinload(Stock.dividends))
    ).all()
    summary = serialize_summary(stocks)
    stock_rows = [serialize_stock(stock) for stock in stocks]

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "汇总"
    summary_sheet.append(["导出时间(UTC)", datetime.now(timezone.utc).isoformat()])
    summary_sheet.append(["基准币种", summary["base_currency"]])
    summary_sheet.append(["持仓总市值", summary["total_market_value"]])
    summary_sheet.append(["上一年总分红", summary["total_latest_full_year_dividend"]])
    summary_sheet.append(["TTM预计总分红", summary["total_annual_dividend"]])
    summary_sheet.append(["上一年组合股息率(%)", summary["overall_latest_full_year_yield"]])
    summary_sheet.append(["TTM组合股息率(%)", summary["overall_dividend_yield"]])
    summary_sheet.append(["股票数量", summary["stock_count"]])

    stocks_sheet = workbook.create_sheet(title="持仓")
    stocks_sheet.append(
        [
            "股票代码",
            "标准代码",
            "股票名称",
            "市场",
            "币种",
            "当前价",
            "上一年",
            "上一年每股分红",
            "上一年股息率(%)",
            "TTM每股分红",
            "TTM股息率(%)",
            "5年平均股息率(%)",
            "10年平均股息率(%)",
            "同步状态",
            "同步时间",
            "同步消息",
        ]
    )
    for item in stock_rows:
        stocks_sheet.append(
            [
                item["symbol"],
                item["normalized_symbol"],
                item["name"],
                item["market"],
                item["currency"],
                item["current_price"],
                item["latest_full_year"],
                item["latest_full_year_dividend"],
                item["latest_full_year_dividend_yield"],
                item["latest_dividend_ttm"],
                item["current_dividend_yield"],
                item["five_year_avg_yield"],
                item["ten_year_avg_yield"],
                item["sync_status"],
                item["last_synced_at"].isoformat() if item["last_synced_at"] else "",
                item["sync_message"] or "",
            ]
        )

    tx_sheet = workbook.create_sheet(title="交易")
    tx_sheet.append(["股票代码", "股票名称", "交易类型", "交易日期", "股数", "均价", "总金额"])
    for item in items:
        for tx in item["transactions"]:
            tx_sheet.append(
                [
                    item["normalized_symbol"],
                    item["name"],
                    tx["transaction_type"],
                    tx["trade_date"].isoformat() if hasattr(tx["trade_date"], "isoformat") else tx["trade_date"],
                    tx["shares"],
                    tx["average_price"],
                    tx["total_amount"],
                ]
            )

    dividend_sheet = workbook.create_sheet(title="分红")
    dividend_sheet.append(["股票代码", "股票名称", "年份", "每股分红", "股息率(%)", "年末价", "币种", "来源"])
    for item in items:
        for dividend in item["dividends"]:
            dividend_sheet.append(
                [
                    item["normalized_symbol"],
                    item["name"],
                    dividend["year"],
                    dividend["dividend_per_share"],
                    dividend["dividend_yield"],
                    dividend["close_price"],
                    dividend["currency"],
                    dividend["source"],
                ]
            )

    content = BytesIO()
    workbook.save(content)
    content.seek(0)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    filename = f"dividend-ledger-export-{ts}.xlsx"
    return StreamingResponse(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/portfolio/import", response_model=PortfolioImportResponse)
def import_portfolio(payload: PortfolioImportRequest, db: Session = Depends(get_db)) -> dict:
    # 导入支持 replace/merge 两种模式，默认 replace 便于整库恢复。
    if payload.mode == "replace":
        db.execute(delete(Transaction))
        db.execute(delete(Dividend))
        db.execute(delete(Stock))
        db.flush()

    imported_stocks = 0
    imported_transactions = 0
    imported_dividends = 0

    for item in payload.stocks:
        # 导入优先使用备份中的符号，保证 export -> import 幂等且不被重新归一化。
        input_symbol = (item.symbol or "").strip().upper()
        input_normalized = (item.normalized_symbol or "").strip().upper()
        resolver_target = input_normalized or input_symbol
        resolved = market_resolver.resolve(resolver_target)
        symbol_value = input_symbol or resolved.normalized_symbol
        normalized_value = input_normalized or resolved.normalized_symbol

        # merge 模式下兼容老数据符号写法差异（例如 BRK_A / BRK-A）。
        stock = db.scalar(
            select(Stock).where(
                or_(
                    Stock.normalized_symbol == normalized_value,
                    Stock.normalized_symbol == resolved.normalized_symbol,
                    Stock.symbol == symbol_value,
                )
            )
        )
        if stock is None:
            stock = Stock(
                symbol=symbol_value,
                normalized_symbol=normalized_value,
                yahoo_symbol=resolved.yahoo_symbol,
                name=item.name.strip(),
                market=resolved.market,
                currency=resolved.currency,
            )
        else:
            stock.symbol = symbol_value
            stock.normalized_symbol = normalized_value
            stock.yahoo_symbol = resolved.yahoo_symbol
            stock.name = item.name.strip()
            stock.market = resolved.market
            stock.currency = resolved.currency
            # merge 场景先落库删除旧明细，再写入新明细，避免唯一约束冲突。
            db.execute(delete(Transaction).where(Transaction.stock_id == stock.id))
            db.execute(delete(Dividend).where(Dividend.stock_id == stock.id))
            db.flush()

        stock.last_price = item.last_price
        stock.latest_dividend_ttm = item.latest_dividend_ttm
        stock.current_dividend_yield = item.current_dividend_yield
        stock.five_year_avg_yield = item.five_year_avg_yield
        stock.ten_year_avg_yield = item.ten_year_avg_yield
        stock.last_synced_at = item.last_synced_at
        stock.sync_status = item.sync_status or "pending"
        stock.sync_message = item.sync_message

        for tx in sorted(item.transactions, key=lambda value: value.trade_date):
            stock.transactions.append(
                Transaction(
                    transaction_type=tx.transaction_type,
                    trade_date=tx.trade_date,
                    shares=shares(tx.shares),
                    average_price=trade_price(tx.average_price),
                    total_amount=money(tx.total_amount or (Decimal(str(tx.shares)) * Decimal(str(tx.average_price)))),
                )
            )
            imported_transactions += 1

        seen_years: set[int] = set()
        for dividend in sorted(item.dividends, key=lambda value: value.year):
            if dividend.year in seen_years:
                continue
            seen_years.add(dividend.year)
            stock.dividends.append(
                Dividend(
                    year=dividend.year,
                    dividend_per_share=dividend.dividend_per_share,
                    dividend_yield=dividend.dividend_yield,
                    close_price=dividend.close_price,
                    currency=dividend.currency or stock.currency,
                    source=dividend.source or "import",
                )
            )
            imported_dividends += 1

        db.add(stock)
        imported_stocks += 1

    db.commit()
    return {
        "imported_stocks": imported_stocks,
        "imported_transactions": imported_transactions,
        "imported_dividends": imported_dividends,
    }


@router.post("/stocks", response_model=StockRowResponse, status_code=status.HTTP_201_CREATED)
def create_stock(payload: CreateStockRequest, background_tasks: BackgroundTasks, db: Session = Depends(get_db)) -> dict:
    # 新建持仓时先落交易记录，再把缓存刷新放到后台执行。
    resolved = market_resolver.resolve(payload.symbol)
    existing = db.scalar(select(Stock).where(Stock.normalized_symbol == resolved.normalized_symbol))
    if existing is not None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="该股票已存在")

    stock = Stock(
        symbol=payload.symbol.strip().upper(),
        normalized_symbol=resolved.normalized_symbol,
        yahoo_symbol=resolved.yahoo_symbol,
        name=payload.name.strip(),
        market=resolved.market,
        currency=resolved.currency,
    )
    db.add(stock)
    db.flush()

    transaction = Transaction(
        stock_id=stock.id,
        transaction_type="buy",
        trade_date=payload.trade_date,
        shares=shares(payload.shares),
        average_price=trade_price(payload.average_price),
        total_amount=money(Decimal(str(payload.shares)) * Decimal(str(payload.average_price))),
    )
    db.add(transaction)
    stock.sync_status = "pending"
    stock.sync_message = "sync scheduled"
    db.commit()

    background_tasks.add_task(sync_stock_cache_task, stock.id)
    stock = load_stock_with_relations(db, stock.id)
    return serialize_stock(stock)


@router.post("/stocks/{stock_id}/transactions", response_model=StockRowResponse)
def create_transaction(
    stock_id: int,
    payload: CreateTransactionRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
) -> dict:
    # 卖出前先校验可卖股数，避免剩余仓位出现负值。
    stock = db.scalar(select(Stock).where(Stock.id == stock_id).options(selectinload(Stock.transactions)))
    if stock is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="股票不存在")

    current_shares, _, _ = summarize_transactions(stock.transactions)
    requested_shares = shares(payload.shares)
    if payload.transaction_type == "sell" and requested_shares > current_shares:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="卖出股数不能超过当前持仓数量")

    transaction = Transaction(
        stock_id=stock.id,
        transaction_type=payload.transaction_type,
        trade_date=payload.trade_date,
        shares=requested_shares,
        average_price=trade_price(payload.average_price),
        total_amount=money(Decimal(str(payload.shares)) * Decimal(str(payload.average_price))),
    )
    db.add(transaction)
    stock.sync_status = "pending"
    stock.sync_message = "sync scheduled"
    db.commit()

    background_tasks.add_task(sync_stock_cache_task, stock.id)
    stock = load_stock_with_relations(db, stock.id)
    return serialize_stock(stock)


@router.get("/stocks/{stock_id}/transactions", response_model=TransactionDetailResponse)
def get_transactions(stock_id: int, db: Session = Depends(get_db)) -> dict:
    stock = load_stock_with_relations(db, stock_id)
    return serialize_transactions(stock)


@router.get("/stocks/{stock_id}/dividends", response_model=DividendDetailResponse)
def get_dividends(stock_id: int, db: Session = Depends(get_db)) -> dict:
    stock = load_stock_with_relations(db, stock_id)
    return serialize_dividends(stock, [])


@router.post("/stocks/{stock_id}/refresh", response_model=StockRowResponse)
def refresh_stock(stock_id: int, db: Session = Depends(get_db)) -> dict:
    try:
        market_data_service.refresh_stock_by_id(db, stock_id)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc)) from exc
    stock = load_stock_with_relations(db, stock_id)
    return serialize_stock(stock)


@router.post("/refresh")
def refresh_all(db: Session = Depends(get_db)) -> dict:
    market_data_service.refresh_all(db)
    return {"status": "ok"}


@router.delete("/stocks/{stock_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_stock(stock_id: int, db: Session = Depends(get_db)) -> None:
    stock = db.get(Stock, stock_id)
    if stock is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="股票不存在")
    db.delete(stock)
    db.commit()
